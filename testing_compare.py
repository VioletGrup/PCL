#!/usr/bin/env python3
from __future__ import annotations

from typing import List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def _read_column_values(
    *,
    excel_path: str,
    sheet_name: str,
    column: str,
    start_row: int,
    end_row: Optional[int] = None,  # ✅ NEW (inclusive)
    max_rows: Optional[int] = None,
) -> List[Tuple[int, object]]:
    """
    Read VALUES ONLY from an Excel column using openpyxl data_only=True.

    Returns a list of (excel_row_number, cell_value).

    Stops when:
      - hits a blank cell (v is None), OR
      - reaches end_row (if provided), OR
      - reaches max_rows (if provided).
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    col_idx = column_index_from_string(column)
    out: List[Tuple[int, object]] = []

    r = start_row
    read_count = 0

    while True:
        # ✅ stop if we have an explicit inclusive end row
        if end_row is not None and r > end_row:
            break

        v = ws.cell(row=r, column=col_idx).value

        # Stop at first empty cell (common spreadsheet convention)
        if v is None:
            break

        out.append((r, v))

        r += 1
        read_count += 1

        if max_rows is not None and read_count >= max_rows:
            break

    return out


def compare_two_excel_columns_values_only(
    *,
    excel_a: str,
    sheet_a: str,
    col_a: str,
    start_row_a: int,
    end_row_a: Optional[int] = None,
    excel_b: str,
    sheet_b: str,
    col_b: str,
    start_row_b: int,
    decimals: int = 12,
    max_rows: Optional[int] = None,
) -> None:
    """
    Compare two columns across two workbooks using computed VALUES (not formulas).
    Floats are compared rounded to `decimals`.
    """
    a = _read_column_values(
        excel_path=excel_a,
        sheet_name=sheet_a,
        column=col_a,
        start_row=start_row_a,
        end_row=end_row_a,
        max_rows=max_rows,
    )
    b = _read_column_values(
        excel_path=excel_b,
        sheet_name=sheet_b,
        column=col_b,
        start_row=start_row_b,
        max_rows=max_rows,
    )

    n = min(len(a), len(b))
    diffs = []

    for i in range(n):
        row_a, va = a[i]
        row_b, vb = b[i]

        # if either is non-numeric, treat as mismatch
        try:
            fa = round(float(va), decimals)
            fb = round(float(vb), decimals)
        except (TypeError, ValueError):
            diffs.append((row_a, va, row_b, vb, "NON_NUMERIC"))
            continue

        if fa != fb:
            if abs(fa) > abs(fb):
                delta = -abs(fa - fb)
            elif abs(fb) > abs(fa):
                delta = abs(fb - fa)
            else:
                delta = 0.0

            diffs.append((row_a, va, row_b, vb, delta))

    # Handle length mismatch
    if len(a) != len(b):
        diffs.append(("LENGTH_MISMATCH", len(a), "vs", len(b), ""))

    if not diffs:
        print(f"MATCH to {decimals} dp.")
        return

    print(
        f"DIFFERENCES found ({decimals} dp):\n"
        f"  A: {excel_a} | {sheet_a}!{col_a} starting row {start_row_a}"
        + (f" ending row {end_row_a}" if end_row_a is not None else "")
        + "\n"
        f"  B: {excel_b} | {sheet_b}!{col_b} starting row {start_row_b}\n"
    )

    for d in diffs:
        if d[0] == "LENGTH_MISMATCH":
            print(f"Length mismatch: A has {d[1]} values, B has {d[3]} values")
            continue

        row_a, va, row_b, vb, delta = d
        if delta == "NON_NUMERIC":
            print(f"A row {row_a}: {va}   |   B row {row_b}: {vb}   (non-numeric)")
        else:
            print(
                f"A row {row_a}: {float(va):.{decimals}f}   |   "
                f"B row {row_b}: {float(vb):.{decimals}f}   |   "
                f"diff = {delta:.{decimals}f}"
            )

    print(f"\nTotal differences: {len(diffs)}")


def compare_results() -> None:
    # -------------------------------
    # EDIT THESE VALUES
    # -------------------------------
    EXCEL_A = "final_pile_elevations_slide_twice.xlsx"
    SHEET_A = "Sheet1"
    COL_A = "H"
    # START_ROW_A = 2  # 1
    # END_ROW_A = 21003  # 1
    # START_ROW_A = 21004  # 2
    # END_ROW_A = 42002  # 2
    START_ROW_A = 42003  # 3
    END_ROW_A = 63008  # 3
    # START_ROW_A = 63009  # 4
    # END_ROW_A = 84401  # 4
    # START_ROW_A = 84402  # 5
    # END_ROW_A = 100411  # 5

    EXCEL_B = "Punchs creek Flat Tracker Imperial-3.xlsm"
    SHEET_B = "Calculations"
    COL_B = "CS"
    START_ROW_B = 9

    compare_two_excel_columns_values_only(
        excel_a=EXCEL_A,
        sheet_a=SHEET_A,
        col_a=COL_A,
        start_row_a=START_ROW_A,
        end_row_a=END_ROW_A,
        excel_b=EXCEL_B,
        sheet_b=SHEET_B,
        col_b=COL_B,
        start_row_b=START_ROW_B,
        decimals=6,
        max_rows=None,  # set e.g. 1000 for testing
    )
