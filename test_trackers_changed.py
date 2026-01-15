#!/usr/bin/env python3
from __future__ import annotations

from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def extract_nonzero_an_values(
    excel_files: Iterable[str],
    sheet_name: str,
    start_row: int = 2,
) -> None:
    """
    For each Excel workbook:
      - read column AN
      - print column A and AN for rows where AN is non-zero

    Parameters
    ----------
    excel_files : Iterable[str]
        Paths to Excel workbooks
    sheet_name : str
        Sheet name to read
    start_row : int
        First Excel row to read (1-based)
    """
    col_a = column_index_from_string("A")
    col_an = column_index_from_string("AN")

    for path in excel_files:
        print("\n" + "=" * 80)
        print(f"Workbook: {path}")
        print("=" * 80)

        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name]

        found = False

        for row in range(start_row, ws.max_row + 1):
            val_an = ws.cell(row=row, column=col_an).value

            # Skip empty cells
            if val_an is None:
                continue

            # Skip zero values
            try:
                if float(val_an) == 0.0:
                    continue
            except (TypeError, ValueError):
                # Non-numeric value → skip
                continue

            val_a = ws.cell(row=row, column=col_a).value

            print(f"Row {row}:  A = {val_a},  AN = {val_an}")
            found = True

        if not found:
            print("No non-zero values found in column AN.")


if __name__ == "__main__":
    # -------------------------------
    # EDIT THESE VALUES
    # -------------------------------
    EXCEL_FILES = [
        "Punchs creek Flat Tracker Imperial-1.xlsm",
        "Punchs creek Flat Tracker Imperial-2.xlsm",
        "Punchs creek Flat Tracker Imperial-3.xlsm",
        "Punchs creek Flat Tracker Imperial-4.xlsm",
        "Punchs creek Flat Tracker Imperial-5.xlsm",
    ]

    SHEET_NAME = "Calculations"
    START_ROW = 9

    extract_nonzero_an_values(
        excel_files=EXCEL_FILES,
        sheet_name=SHEET_NAME,
        start_row=START_ROW,
    )
