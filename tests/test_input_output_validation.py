#!/usr/bin/env python3
"""
Get data
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from BasePile import BasePile
from BaseTracker import BaseTracker
from flatTrackerGrading import main
from Project import Project
from ProjectConstraints import ProjectConstraints


def load_project_from_excel(
    *,
    excel_path: str,
    sheet_name: str,
    project_name: str,
    project_type: str,
    constraints: ProjectConstraints,
) -> Project:
    """
    Initialise a Project from an Excel sheet containing tracker + pile rows.

    Expected Excel columns (1-indexed as you described):
      - Col 1: tracker number
      - Col 3: pile in tracker
      - Col 4: easting
      - Col 5: northing
      - Col 9: ground elevation (initial_elevation)

    Pile ID is constructed as "tracker_number.pile_in_tracker" (string).
    Flooding allowance is set to 0 for all piles.
    """
    # Read raw sheet (no header assumptions)
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)

    # Drop rows where tracker number is not numeric
    df = df[pd.to_numeric(df.iloc[:, 0], errors="coerce").notna()]

    # If your sheet has headers, we still use positional indexing by column number.
    # Convert to 0-based indexes:
    # col1 becomes 0, col3 becomes 2, col4 becomes 3, col5 becomes 4, col9 becomes 8
    tracker_col = df.columns[0]
    pile_in_tracker_col = df.columns[2]
    easting_col = df.columns[3]
    northing_col = df.columns[4]
    elevation_col = df.columns[8]

    # Drop rows missing essential fields (common with blank separators)
    df = df.dropna(
        subset=[tracker_col, pile_in_tracker_col, easting_col, northing_col, elevation_col]
    )

    project = Project(
        name=project_name,
        project_type=project_type,  # e.g. "standard"
        constraints=constraints,
    )

    trackers_by_id: Dict[int, BaseTracker] = {}
    piles = 0

    for _, row in df.iterrows():
        piles += 1
        tracker_num = int(row[tracker_col])
        pit = int(row[pile_in_tracker_col])

        easting = float(row[easting_col])
        northing = float(row[northing_col])
        ground_z = float(row[elevation_col])

        # Create tracker if needed
        if tracker_num not in trackers_by_id:
            trackers_by_id[tracker_num] = BaseTracker(tracker_id=tracker_num)

        # pile_id = float(f"{tracker_num}.{pit}")
        pile_id = float(f"{tracker_num}.{pit:02d}")

        trackers_by_id[tracker_num].piles.append(
            BasePile(
                northing=northing,
                easting=easting,
                initial_elevation=ground_z,
                pile_id=pile_id,  # NOTE: string id
                pile_in_tracker=pit,
                flooding_allowance=0.0,
            )
        )

    # Attach trackers in numeric order and sort their piles
    for tid in sorted(trackers_by_id.keys()):
        t = trackers_by_id[tid]
        t.sort_by_pole_position()
        project.trackers.append(t)

    return project


def to_excel(project: Project) -> None:
    rows = []

    for tracker in project.trackers:
        for pile in tracker.piles:
            rows.append(
                {
                    "tracker_id": tracker.tracker_id,
                    "pile_id": pile.pile_id,
                    "pile_in_tracker": pile.pile_in_tracker,
                    "northing": pile.northing,
                    "easting": pile.easting,
                    "initial_elevation": pile.initial_elevation,
                    "final_elevation": pile.final_elevation,
                    "change": pile.final_elevation - pile.initial_elevation,
                    "total_height": pile.total_height,
                    "total_revealed": pile.pile_revealed,
                }
            )

    df = pd.DataFrame(rows)

    output_path = "final_pile_elevations.xlsx"
    df.to_excel(output_path, index=False)


def _read_column_values(
    *,
    excel_path: str,
    sheet_name: str,
    column: str,
    start_row: int,
    max_rows: Optional[int] = None,
) -> List[Tuple[int, object]]:
    """
    Read VALUES ONLY from an Excel column using openpyxl data_only=True.

    Returns a list of (excel_row_number, cell_value).
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    col_idx = column_index_from_string(column)
    out: List[Tuple[int, object]] = []

    r = start_row
    read_count = 0

    while True:
        v = ws.cell(row=r, column=col_idx).value

        # Stop at first empty cell (common spreadsheet convention)
        if v is None:
            break

        out.append((r, v))

        r += 1
        read_count += 1

        if max_rows is not None and read_count >= max_rows:
            break

    return out


def compare_two_excel_columns_values_only(
    *,
    excel_a: str,
    sheet_a: str,
    col_a: str,
    start_row_a: int,
    excel_b: str,
    sheet_b: str,
    col_b: str,
    start_row_b: int,
    decimals: int = 6,
    max_rows: Optional[int] = None,
) -> List[str]:
    """
    Compare two columns across two workbooks using computed VALUES (not formulas).
    Floats are compared rounded to `decimals`.
    Returns list of difference strings (empty if match).
    """
    a = _read_column_values(
        excel_path=excel_a,
        sheet_name=sheet_a,
        column=col_a,
        start_row=start_row_a,
        max_rows=max_rows,
    )
    b = _read_column_values(
        excel_path=excel_b,
        sheet_name=sheet_b,
        column=col_b,
        start_row=start_row_b,
        max_rows=max_rows,
    )

    n = min(len(a), len(b))
    diffs = []

    for i in range(n):
        row_a, va = a[i]
        row_b, vb = b[i]

        # if either is non numeric, treat as mismatch
        try:
            fa = round(float(va), decimals)
            fb = round(float(vb), decimals)
        except (TypeError, ValueError):
            diffs.append(f"A row {row_a}: {va}   |   B row {row_b}: {vb}   (non-numeric)")
            continue

        if fa != fb:
            diffs.append(
                f"A row {row_a}: {float(va):.{decimals}f}   |   "
                f"B row {row_b}: {float(vb):.{decimals}f}   |   "
                f"diff = {fa - fb:.{decimals}f}"
            )

    # Handle length mismatch
    if len(a) != len(b):
        diffs.append(f"LENGTH_MISMATCH: A has {len(a)} values, B has {len(b)} values")

    return diffs


class TestDataIO:

    def test_load_project_works(self):
        """Test that load_project_from_excel works (testing_get_data.py logic)."""
        root_dir = Path(__file__).parent.parent
        input_excel = str(root_dir / "Test Piling Info.xlsx")

        constraints = ProjectConstraints(
            min_reveal_height=1.375,
            max_reveal_height=1.675,
            pile_install_tolerance=0.0,
            max_incline=0.15,
            target_height_percentage=0.5,
            max_angle_rotation=0.0,
            edge_overhang=0.0,
        )

        project = load_project_from_excel(
            excel_path=input_excel,
            sheet_name="Piling information",
            project_name="Test_Load",
            project_type="standard",
            constraints=constraints,
        )

        # Verify loaded
        assert project is not None
        assert len(project.trackers) > 0
        assert project.total_piles > 0

    def test_export_creates_file(self, tmp_path):
        """Test that to_excel creates the output file (testing_get_data.py logic)."""
        import os

        root_dir = Path(__file__).parent.parent
        input_excel = str(root_dir / "Test Piling Info.xlsx")

        constraints = ProjectConstraints(
            min_reveal_height=1.375,
            max_reveal_height=1.675,
            pile_install_tolerance=0.0,
            max_incline=0.15,
            target_height_percentage=0.5,
            max_angle_rotation=0.0,
            edge_overhang=0.0,
        )

        project = load_project_from_excel(
            excel_path=input_excel,
            sheet_name="Piling information",
            project_name="Test_Export",
            project_type="standard",
            constraints=constraints,
        )

        main(project)

        original_dir = os.getcwd()
        os.chdir(tmp_path)

        try:
            to_excel(project)
            output_file = tmp_path / "final_pile_elevations.xlsx"
            assert output_file.exists()

        finally:
            os.chdir(original_dir)

    def test_compare_function_detects_differences(self, tmp_path):
        """Test that compare function correctly detects differences."""
        # Create two Excel files with different values
        df_a = pd.DataFrame({"values": [1.123456, 2.234567, 3.345678]})
        df_b = pd.DataFrame({"values": [1.123457, 2.234567, 3.345678]})  # First differs

        file_a = tmp_path / "test_a.xlsx"
        file_b = tmp_path / "test_b.xlsx"

        df_a.to_excel(file_a, index=False)
        df_b.to_excel(file_b, index=False)

        diffs = compare_two_excel_columns_values_only(
            excel_a=str(file_a),
            sheet_a="Sheet1",
            col_a="A",
            start_row_a=2,  # Skip header
            excel_b=str(file_b),
            sheet_b="Sheet1",
            col_b="A",
            start_row_b=2,
            decimals=6,
        )

        # Should find 1 difference
        assert len(diffs) == 1
        assert "1.123456" in diffs[0] or "1.123457" in diffs[0]

    def test_compare_function_finds_match(self, tmp_path):
        """Test that compare function returns empty list when values match."""
        # Create two identical Excel files
        df = pd.DataFrame({"values": [1.123456, 2.234567, 3.345678]})

        file_a = tmp_path / "test_a.xlsx"
        file_b = tmp_path / "test_b.xlsx"

        df.to_excel(file_a, index=False)
        df.to_excel(file_b, index=False)

        diffs = compare_two_excel_columns_values_only(
            excel_a=str(file_a),
            sheet_a="Sheet1",
            col_a="A",
            start_row_a=2,
            excel_b=str(file_b),
            sheet_b="Sheet1",
            col_b="A",
            start_row_b=2,
            decimals=6,
        )

        # Should find no differences
        assert len(diffs) == 0
